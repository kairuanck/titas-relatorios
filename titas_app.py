"""
TITÃS — GERADOR DE RELATÓRIOS 2026
Interface web com upload de arquivos

Dependências:  pip install streamlit pandas openpyxl
Uso:           streamlit run titas_app.py
"""

import io
from datetime import datetime
import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Configuração ───────────────────────────────────────────────────────────
st.set_page_config(page_title="Titãs · Relatórios", page_icon="📊", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Sans:wght@300;400;500&family=DM+Mono&display=swap');

html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
.block-container { padding: 2rem 3rem; max-width: 1200px; }

.hero {
    background: linear-gradient(135deg, #0f1f3d 0%, #1a3a6b 60%, #0e4d8c 100%);
    border-radius: 20px; padding: 3rem; color: white;
    margin-bottom: 2.5rem; position: relative; overflow: hidden;
}
.hero::before {
    content: ''; position: absolute; top: -60px; right: -60px;
    width: 260px; height: 260px; border-radius: 50%;
    background: rgba(255,255,255,0.04); pointer-events: none;
}
.hero::after {
    content: ''; position: absolute; bottom: -40px; right: 120px;
    width: 140px; height: 140px; border-radius: 50%;
    background: rgba(255,255,255,0.03); pointer-events: none;
}
.hero h1 {
    font-family: 'Syne', sans-serif; font-size: 2.6rem;
    font-weight: 800; margin: 0 0 .5rem; letter-spacing: -.5px;
}
.hero p { font-size: .95rem; opacity: .75; margin: 0; font-weight: 300; }

.section-label {
    font-family: 'Syne', sans-serif; font-weight: 700;
    font-size: .7rem; letter-spacing: .12em; text-transform: uppercase;
    color: #2471A3; margin-bottom: .5rem;
}
.upload-card {
    border: 1.5px dashed #C8D8EC; border-radius: 14px;
    padding: 1.2rem 1.4rem; background: #F7FAFF;
    margin-bottom: .8rem; transition: border-color .2s;
}
.upload-card:hover { border-color: #2471A3; }
.upload-card h4 {
    font-family: 'Syne', sans-serif; font-weight: 700;
    font-size: .9rem; color: #1F3864; margin: 0 0 .2rem;
}
.upload-card p { font-size: .78rem; color: #7A90A8; margin: 0; }

.metric-row { display: flex; gap: 12px; margin: 1.5rem 0; flex-wrap: wrap; }
.metric-box {
    flex: 1; min-width: 140px; background: white;
    border: 1px solid #E0EAF6; border-radius: 12px;
    padding: 1rem 1.2rem; text-align: center;
}
.metric-label {
    font-size: .7rem; font-weight: 600; letter-spacing: .08em;
    text-transform: uppercase; color: #7A90A8;
}
.metric-value {
    font-family: 'DM Mono', monospace; font-size: 1.4rem;
    font-weight: 500; color: #1F3864; margin: .3rem 0 .1rem;
}
.metric-sub { font-size: .75rem; color: #9AAABB; }

.status-ok   { color: #1A6B3A; background: #D5F5E3; padding: 2px 10px; border-radius: 99px; font-size: .78rem; font-weight: 600; }
.status-warn { color: #7D5A00; background: #FFF3CD; padding: 2px 10px; border-radius: 99px; font-size: .78rem; font-weight: 600; }
.status-bad  { color: #922B21; background: #FADBD8; padding: 2px 10px; border-radius: 99px; font-size: .78rem; font-weight: 600; }

.stDownloadButton > button {
    width: 100%; background: linear-gradient(135deg,#1F3864,#2471A3);
    color: white; border: none; border-radius: 10px;
    padding: .65rem 1.5rem; font-weight: 600; font-size: .9rem;
    letter-spacing: .02em;
}
.stDownloadButton > button:hover { opacity: .9; }

.run-btn > button {
    background: linear-gradient(135deg, #1a6b3a, #27ae60) !important;
    color: white !important; border: none !important;
    border-radius: 12px !important; font-size: 1.05rem !important;
    font-weight: 700 !important; padding: .8rem 2rem !important;
    width: 100% !important; letter-spacing: .03em !important;
}
</style>
""", unsafe_allow_html=True)

# ── Constantes ─────────────────────────────────────────────────────────────
UF_COL_TO_LAB = {
    "RS": {
        "AGROLIFE":"AGROLIFE RS","AVERT":"AVERT RS","BBPET":"MUNDO ANIMAL RS",
        "BBPET ESTETICA":"MA ESTÉTICA RS","BIOCLIN":"BIOCLIN","BRINDE":"BRINDE",
        "CEVA":"CEVA","OURO FINO":"OUROFINO RS","OURO FINO WELLPET":"WELLPET",
        "SPIN":"SPIN RS","SYNTEC":"SYNTEC","WANPY":"WANPY",
    },
    "RJ": {"AGROLIFE":"AGROLIFE RJ","AVERT":"AVERT RJ","ELANCO":"ELANCO RJ"},
    "PR": {"BBPET":"MUNDO ANIMAL PR","ELANCO":"ELANCO PR","SPIN":"SPIN PR"},
}

LAB_ORDER = [
    "CEVA","SYNTEC","MUNDO ANIMAL RS","MA ESTÉTICA RS","BIOCLIN","WANPY",
    "OUROFINO RS","WELLPET","AGROLIFE RS","SPIN RS","AVERT RS",
    "AVERT RJ","ELANCO RJ","AGROLIFE RJ",
    "ELANCO PR","AGROLIFE PR","MUNDO ANIMAL PR","MA ESTÉTICA PR","WANPY PR","SPIN PR",
    "BRINDE",
]

MESES = ["JANEIRO","FEVEREIRO","MARÇO","ABRIL","MAIO","JUNHO",
         "JULHO","AGOSTO","SETEMBRO","OUTUBRO","NOVEMBRO","DEZEMBRO"]
VENDEDORAS_ORDER = ["Emily","Rosiris","Lizie","Tamires","Claudia"]
VC_HEX = {"Emily":"154360","Rosiris":"1A5276","Lizie":"1F618D",
           "Tamires":"2471A3","Claudia":"2980B9"}

# ── Excel helpers ──────────────────────────────────────────────────────────
def tb():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def fill(h): return PatternFill("solid", start_color=h, end_color=h)

HDR=fill("1F3864"); SUB=fill("2E75B6"); GOLD=fill("FFD700")
ALT=fill("EBF5FB"); POS=fill("D5F5E3"); NEG=fill("FADBD8"); GRY=fill("F2F3F4")

def sty(cell, value=None, bold=False, size=10, color="000000",
        bg=None, h="center", wrap=False):
    if value is not None: cell.value = value
    cell.font = Font(name="Arial", bold=bold, size=size, color=color)
    if bg: cell.fill = bg
    cell.alignment = Alignment(horizontal=h, vertical="center", wrap_text=wrap)
    cell.border = tb()

def money(cell, value=None, bold=False, bg=None, color="000000"):
    if value is not None: cell.value = value
    cell.number_format = "R$ #,##0.00"
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.font = Font(name="Arial", bold=bold, size=10, color=color)
    cell.border = tb()
    if bg: cell.fill = bg

def pct_cell(cell, value=None, bg=None, color="000000", bold=False):
    if value is not None: cell.value = value
    cell.number_format = "0.0%"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(name="Arial", size=10, bold=bold, color=color)
    cell.border = tb()
    if bg: cell.fill = bg

def title_row(ws, text, n, row=1, bg=None):
    ws.merge_cells(f"A{row}:{get_column_letter(n)}{row}")
    sty(ws.cell(row=row, column=1), text, bold=True, size=13,
        color="FFFFFF", bg=bg or HDR)
    ws.row_dimensions[row].height = 32

def header_row(ws, headers, row=2, bg=None):
    for i, h in enumerate(headers, 1):
        sty(ws.cell(row=row, column=i), h, bold=True, size=10,
            color="FFFFFF", bg=bg or SUB, wrap=True)
    ws.row_dimensions[row].height = 26

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def sum_col(ws, row, ci, ds, bold=True, bg=None, color="000000"):
    c = ws.cell(row=row, column=ci)
    c.value = f"=SUM({get_column_letter(ci)}{ds}:{get_column_letter(ci)}{row-1})"
    c.number_format = "R$ #,##0.00"
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.font = Font(name="Arial", bold=bold, size=10, color=color)
    c.border = tb()
    if bg: c.fill = bg

# ── Leitura de dados ────────────────────────────────────────────────────────
def load_base(file_bytes):
    sheets = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None)
    parts = [df.assign(Vendedora=v) for v, df in sheets.items()]
    base = pd.concat(parts, ignore_index=True)
    base["Código"] = base["Código"].astype(str).str.strip()
    return base, list(sheets.keys())

def load_meta(file_bytes):
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, header=None)
    nome_map = {"EMILY":"Emily","ROSIRIS":"Rosiris","LIZIE":"Lizie",
                "TAMIRES":"Tamires","CLAUDIA":"Claudia"}
    sections = []
    for i, row in df.iterrows():
        for cell in row:
            if isinstance(cell, str):
                first = cell.strip().upper().split()[0] if cell.strip() else ""
                if first in nome_map:
                    sections.append((i, nome_map[first])); break
    metas = {}
    for idx, (start, vendedora) in enumerate(sections):
        end = sections[idx+1][0] if idx+1 < len(sections) else len(df)
        hdr_idx = None
        for r in range(start, min(start+5, end)):
            vals = [str(v).strip().upper() for v in df.iloc[r] if pd.notna(v)]
            if "LAB" in vals and "META" in vals:
                hdr_idx = r; break
        if hdr_idx is None: continue
        hvals = [str(v).strip().upper() if pd.notna(v) else "" for v in df.iloc[hdr_idx]]
        try:
            col_lab = hvals.index("LAB"); col_meta = hvals.index("META")
        except ValueError: continue
        col_fat = hvals.index("FATURADO") if "FATURADO" in hvals else None
        labs = {}
        for r in range(hdr_idx+1, end):
            lv = df.iat[r, col_lab]; mv = df.iat[r, col_meta]
            if not isinstance(lv, str) or pd.isna(lv): continue
            lab = lv.strip()
            if lab in ("LAB","TOTAL",""): continue
            if not isinstance(mv, (int, float)) or pd.isna(mv): continue
            fat = None
            if col_fat is not None:
                fv = df.iat[r, col_fat]
                fat = float(fv) if isinstance(fv,(int,float)) and not pd.isna(fv) else None
            labs[lab] = {"meta": float(mv), "faturado": fat}
        if vendedora not in metas:
            metas[vendedora] = labs
    return metas

def load_data_file(file_bytes, uf):
    raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name="Export", header=None)
    h = [str(v).strip() if pd.notna(v) else "" for v in raw.iloc[0]]
    brands = [v for v in h if v and v not in ("Departamento","Total","")]
    df = raw.iloc[2:].copy()
    dynamic_cols = ["Cliente","Cidade"] + brands
    if h[-1].upper() in ("TOTAL","FAT.LIQ. R$",""):
        dynamic_cols.append("Total")
    n_file, n_exp = len(df.columns), len(dynamic_cols)
    if n_file == n_exp: df.columns = dynamic_cols
    elif n_file < n_exp: df.columns = dynamic_cols[:n_file]
    else: df.columns = dynamic_cols + [f"_x{i}" for i in range(n_file-n_exp)]

    df = df[~df["Cliente"].astype(str).str.startswith("Filtros")]
    df = df.dropna(subset=["Cliente"])
    df["Código"] = df["Cliente"].astype(str).str.extract(r"^(\d+)")
    df = df.dropna(subset=["Código"])
    df["Nome Cliente"] = df["Cliente"].str.replace(r"^\d+\s*-\s*","",regex=True)
    for c in brands + (["Total"] if "Total" in df.columns else []):
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    mapping = UF_COL_TO_LAB.get(uf, {})
    result = df[["Código","Nome Cliente","Cidade"]].copy()
    for brand, lab in mapping.items():
        if brand not in df.columns: continue
        if lab not in result.columns: result[lab] = 0.0
        result[lab] = result.get(lab, pd.Series(0.0, index=result.index)) + df[brand].fillna(0)

    lab_cols = [lab for lab in mapping.values() if lab in result.columns]
    result = result.groupby("Código").agg(
        **{"Nome Cliente":("Nome Cliente","first"),
           "Cidade":("Cidade","first"),
           **{lab:(lab,"sum") for lab in lab_cols}}
    )
    return result

def load_all_data(files_uf, base):
    """files_uf: dict {UF: bytes}"""
    result = None
    for uf in ["RS","RJ","PR"]:
        if uf not in files_uf: continue
        df_uf = load_data_file(files_uf[uf], uf)
        if result is None:
            result = df_uf
        else:
            lab_cols = [c for c in df_uf.columns if c not in ("Nome Cliente","Cidade")]
            result = result.join(df_uf[lab_cols], how="outer")

    if result is None: return None
    for lab in LAB_ORDER:
        if lab not in result.columns: result[lab] = 0.0
    result = result.reset_index()

    base_info = (base[["Código","Vendedora","Estado","Filial de Faturamento"]]
                 .drop_duplicates("Código").set_index("Código"))
    result = result.set_index("Código").join(base_info, how="left").reset_index()
    result["Vendedora"] = result["Vendedora"].fillna("NÃO IDENTIFICADA")
    result["Estado"]    = result["Estado"].fillna("")

    labs_present = [l for l in LAB_ORDER if l in result.columns]
    result["Total"] = result[labs_present].sum(axis=1)
    return result

# ── Geradores de Excel ─────────────────────────────────────────────────────
def gerar_cruzamento(combined, mes):
    df_s = combined[combined["Total"] > 0].copy()
    labs = [l for l in LAB_ORDER if l in df_s.columns and df_s[l].sum() > 0]
    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "Cruzamento de Vendas"; ws1.freeze_panes = "A3"
    h1 = ["Vendedora","Código","Nome Cliente","Cidade","Estado","Filial"] + labs + ["TOTAL"]
    title_row(ws1, f"CRUZAMENTO BASE TITÃS × VENDAS — {mes} 2026", len(h1))
    header_row(ws1, h1)
    row = 3
    for _, r in df_s.sort_values(["Vendedora","Total"], ascending=[True,False]).iterrows():
        is_nid = r["Vendedora"] == "NÃO IDENTIFICADA"
        bg = fill("FFCCCC") if is_nid else (ALT if row%2==0 else None)
        for ci, v in enumerate([r["Vendedora"],r["Código"],r.get("Nome Cliente",""),
                                  r.get("Cidade",""),r.get("Estado",""),
                                  r.get("Filial de Faturamento","")], 1):
            sty(ws1.cell(row=row,column=ci), str(v) if pd.notna(v) else "", bg=bg, h="left")
        for ci, lab in enumerate(labs+["Total"], 7):
            v = float(r.get(lab, 0))
            money(ws1.cell(row=row,column=ci), v if v>0 else None, bg=bg)
        row += 1
    for ci in range(1, len(h1)+1):
        c = ws1.cell(row=row, column=ci); c.fill = GOLD; c.border = tb()
        if ci == 1: sty(c, "TOTAL GERAL", bold=True, bg=GOLD)
        elif ci > 6: sum_col(ws1, row, ci, 3, bg=GOLD, color="000000")
    set_widths(ws1, [16,8,32,18,7,7]+[13]*len(labs)+[14])

    ws2 = wb.create_sheet("Resumo por Vendedora"); ws2.freeze_panes = "A3"
    h2 = ["Vendedora","Qtd Clientes"]+labs+["TOTAL GERAL"]
    title_row(ws2, f"RESUMO DE VENDAS POR VENDEDORA — {mes} 2026", len(h2))
    header_row(ws2, h2)
    grp = df_s.groupby("Vendedora")[labs+["Total"]].sum().reset_index()
    qtd = df_s.groupby("Vendedora")["Código"].count().reset_index(name="Qtd")
    grp = grp.merge(qtd, on="Vendedora")
    row2 = 3
    for _, r in grp.sort_values("Vendedora").iterrows():
        is_nid = r["Vendedora"] == "NÃO IDENTIFICADA"
        bg = fill("FFCCCC") if is_nid else (ALT if row2%2==0 else None)
        sty(ws2.cell(row=row2,column=1), r["Vendedora"], bg=bg, h="left")
        sty(ws2.cell(row=row2,column=2), int(r["Qtd"]), bg=bg)
        for ci, col in enumerate(labs+["Total"], 3):
            v = float(r.get(col, 0))
            money(ws2.cell(row=row2,column=ci), v if v>0 else None, bg=bg)
        row2 += 1
    for ci in range(1, len(h2)+1):
        c = ws2.cell(row=row2, column=ci); c.fill = GOLD; c.border = tb()
        if ci == 1: sty(c, "TOTAL GERAL", bold=True, bg=GOLD)
        elif ci > 2: sum_col(ws2, row2, ci, 3, bg=GOLD, color="000000")
    set_widths(ws2, [18,14]+[13]*len(labs)+[14])

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, len(df_s)

def gerar_meta_vs_real(metas, combined, mes):
    vendedoras = [v for v in VENDEDORAS_ORDER if v in metas]
    merged = combined.copy()
    merged["Vendedora"] = merged["Vendedora"].fillna("NÃO IDENTIFICADA")
    all_labs = [l for l in LAB_ORDER
                if l in combined.columns or any(l in m for m in metas.values())]
    for lab in all_labs:
        if lab not in merged.columns: merged[lab] = 0.0
    realizado = merged.groupby("Vendedora")[all_labs].sum()

    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "Meta vs Realizado"; ws1.freeze_panes = "B4"
    CPV = 4; N = 1 + len(vendedoras)*CPV + CPV
    title_row(ws1, f"META VS. REALIZADO — {mes} 2026 | TITÃS", N)

    ws1.cell(row=2,column=1).fill = HDR; ws1.cell(row=2,column=1).border = tb()
    col = 2
    for v in vendedoras:
        ws1.merge_cells(f"{get_column_letter(col)}2:{get_column_letter(col+3)}2")
        sty(ws1.cell(row=2,column=col), v.upper(), bold=True, size=11,
            color="FFFFFF", bg=fill(VC_HEX[v]))
        for x in range(1,4):
            ws1.cell(row=2,column=col+x).fill = fill(VC_HEX[v])
            ws1.cell(row=2,column=col+x).border = tb()
        col += 4
    ws1.merge_cells(f"{get_column_letter(col)}2:{get_column_letter(col+3)}2")
    sty(ws1.cell(row=2,column=col), "TOTAL GERAL", bold=True, size=11,
        color="FFFFFF", bg=HDR)
    for x in range(1,4):
        ws1.cell(row=2,column=col+x).fill = HDR
        ws1.cell(row=2,column=col+x).border = tb()
    ws1.row_dimensions[2].height = 26

    sty(ws1.cell(row=3,column=1), "LABORATÓRIO / DEPTO.", bold=True, color="FFFFFF", bg=SUB)
    col = 2
    for _ in range(len(vendedoras)+1):
        for lbl in ["META","REALIZADO","GAP","% ATING."]:
            sty(ws1.cell(row=3,column=col), lbl, bold=True, size=9,
                color="FFFFFF", bg=SUB, wrap=True); col += 1
    ws1.row_dimensions[3].height = 28

    DS = 4; row = DS
    for lab in all_labs:
        sem = (lab == "BRINDE")
        base_bg = GRY if sem else (ALT if row%2==0 else None)
        sty(ws1.cell(row=row,column=1), lab, h="left", bg=base_bg)
        col = 2; t_m = t_r = 0
        for v in vendedoras:
            m_val = metas.get(v,{}).get(lab,{}).get("meta",0.0) if not sem else 0.0
            r_val = float(realizado.loc[v,lab]) if v in realizado.index else 0.0
            g = r_val-m_val; p = r_val/m_val if m_val else None
            t_m += m_val; t_r += r_val
            gf = GRY if sem else (POS if g>=0 and m_val>0 else (NEG if m_val>0 else base_bg))
            gc = ("1A5E1A" if not sem and g>=0 and m_val>0
                  else "CC0000" if not sem and g<0 and m_val>0 else "000000")
            money(ws1.cell(row=row,column=col),   m_val if m_val else None, bg=base_bg)
            money(ws1.cell(row=row,column=col+1), r_val if r_val else None, bg=base_bg)
            money(ws1.cell(row=row,column=col+2), g if (m_val or r_val) else None, bg=gf, color=gc)
            pct_cell(ws1.cell(row=row,column=col+3), p, bg=gf); col += 4
        tg = t_r-t_m; tp = t_r/t_m if t_m else None
        tf = GRY if sem else (POS if tg>=0 and t_m>0 else NEG)
        tc = ("1A5E1A" if not sem and tg>=0 and t_m>0
              else "CC0000" if not sem and tg<0 else "000000")
        money(ws1.cell(row=row,column=col),   t_m if t_m else None, bold=True, bg=base_bg)
        money(ws1.cell(row=row,column=col+1), t_r if t_r else None, bold=True, bg=base_bg)
        money(ws1.cell(row=row,column=col+2), tg if (t_m or t_r) else None, bold=True, bg=tf, color=tc)
        pct_cell(ws1.cell(row=row,column=col+3), tp, bg=tf, bold=True); row += 1

    tr = row
    sty(ws1.cell(row=tr,column=1), "TOTAL GERAL", bold=True, size=11, color="FFFFFF", bg=HDR)
    col = 2; gm = gr_tot = 0; summary = {}
    for v in vendedoras:
        vm = sum(d["meta"] for lab,d in metas.get(v,{}).items() if lab != "BRINDE")
        vr = float(realizado.loc[v,all_labs].sum()) if v in realizado.index else 0.0
        vg = vr-vm; vp = vr/vm if vm else None
        gm += vm; gr_tot += vr
        summary[v] = {"meta":vm,"real":vr,"gap":vg,"pct":vp or 0}
        money(ws1.cell(row=tr,column=col),   vm, bold=True, bg=HDR, color="FFFFFF")
        money(ws1.cell(row=tr,column=col+1), vr, bold=True, bg=HDR, color="FFFFFF")
        money(ws1.cell(row=tr,column=col+2), vg, bold=True, bg=HDR, color="FFFFFF")
        pct_cell(ws1.cell(row=tr,column=col+3), vp, bg=HDR, color="FFFFFF", bold=True); col += 4
    gg = gr_tot-gm; gp = gr_tot/gm if gm else None
    money(ws1.cell(row=tr,column=col),   gm, bold=True, bg=HDR, color="FFFFFF")
    money(ws1.cell(row=tr,column=col+1), gr_tot, bold=True, bg=HDR, color="FFFFFF")
    money(ws1.cell(row=tr,column=col+2), gg, bold=True, bg=HDR, color="FFFFFF")
    pct_cell(ws1.cell(row=tr,column=col+3), gp, bg=HDR, color="FFFFFF", bold=True)
    ws1.row_dimensions[tr].height = 22
    ws1.column_dimensions["A"].width = 26
    col = 2
    for _ in range(len(vendedoras)+1):
        for w in [13,13,13,9]:
            ws1.column_dimensions[get_column_letter(col)].width = w; col += 1

    ws2 = wb.create_sheet("Resumo por Vendedora"); ws2.freeze_panes = "A3"
    title_row(ws2, f"RESUMO META VS. REALIZADO — {mes} 2026", 6)
    header_row(ws2, ["VENDEDORA","META TOTAL","REALIZADO","GAP","% ATINGIMENTO","STATUS"])
    row2 = 3
    for v in vendedoras:
        s = summary[v]; p = s["pct"]
        sf = POS if p>=1.0 else (ALT if p>=0.85 else NEG)
        status = "✅ ACIMA DA META" if p>=1.0 else ("⚠️ PRÓXIMO" if p>=0.85 else "❌ ABAIXO DA META")
        sty(ws2.cell(row=row2,column=1), v.upper(), bold=True, size=11,
            color="FFFFFF", bg=fill(VC_HEX[v]))
        money(ws2.cell(row=row2,column=2), s["meta"], bg=sf)
        money(ws2.cell(row=row2,column=3), s["real"], bg=sf)
        money(ws2.cell(row=row2,column=4), s["gap"], bg=sf,
              color="CC0000" if s["gap"]<0 else "1A5E1A")
        pct_cell(ws2.cell(row=row2,column=5), p, bg=sf)
        sty(ws2.cell(row=row2,column=6), status, bg=sf)
        ws2.row_dimensions[row2].height = 22; row2 += 1
    gp2 = gr_tot/gm if gm else 0; gg2 = gr_tot-gm
    gs = "✅ ACIMA DA META" if gp2>=1.0 else ("⚠️ PRÓXIMO" if gp2>=0.85 else "❌ ABAIXO DA META")
    sty(ws2.cell(row=row2,column=1), "TOTAL TITÃS", bold=True, size=11, color="FFFFFF", bg=HDR)
    money(ws2.cell(row=row2,column=2), gm, bold=True, bg=HDR, color="FFFFFF")
    money(ws2.cell(row=row2,column=3), gr_tot, bold=True, bg=HDR, color="FFFFFF")
    money(ws2.cell(row=row2,column=4), gg2, bold=True, bg=HDR, color="FFFFFF")
    pct_cell(ws2.cell(row=row2,column=5), gp2, bg=HDR, color="FFFFFF", bold=True)
    sty(ws2.cell(row=row2,column=6), gs, bold=True, color="FFFFFF", bg=HDR)
    ws2.row_dimensions[row2].height = 24
    set_widths(ws2, [18,16,16,16,15,20])

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, summary, gm, gr_tot

def gerar_metas_excel(metas, mes):
    vendedoras = [v for v in VENDEDORAS_ORDER if v in metas]
    all_labs_set = {lab for v in vendedoras for lab in metas[v]}
    ordered = [l for l in LAB_ORDER if l in all_labs_set] + \
              [l for l in all_labs_set if l not in LAB_ORDER]
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Resumo por Laboratório"; ws.freeze_panes = "B3"
    N = 1+len(vendedoras)+1
    title_row(ws, f"METAS {mes} 2026 — COMPILADO POR LABORATÓRIO", N)
    header_row(ws, ["LABORATÓRIO"]+vendedoras+["TOTAL GERAL"])
    DS = 3
    for i, lab in enumerate(ordered, DS):
        bg = ALT if i%2==0 else None
        sty(ws.cell(row=i,column=1), lab, bg=bg, h="left")
        for ci, v in enumerate(vendedoras, 2):
            val = metas.get(v,{}).get(lab,{}).get("meta",None)
            money(ws.cell(row=i,column=ci), val if val else None, bg=bg)
        lv = get_column_letter(1+len(vendedoras))
        c = ws.cell(row=i, column=N)
        c.value = f"=SUM(B{i}:{lv}{i})"
        c.number_format = "R$ #,##0.00"
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.font = Font(name="Arial", bold=True); c.border = tb()
        if bg: c.fill = bg
    tr = DS+len(ordered)
    for ci in range(1, N+1):
        c = ws.cell(row=tr, column=ci); c.fill = HDR; c.border = tb()
        if ci == 1: sty(c, "TOTAL GERAL", bold=True, size=11, color="FFFFFF", bg=HDR)
        else: sum_col(ws, tr, ci, DS)
    set_widths(ws, [26]+[16]*(len(vendedoras)+1))
    for v in vendedoras:
        wv = wb.create_sheet(v); wv.freeze_panes = "B3"
        title_row(wv, f"METAS {mes} 2026 — {v.upper()}", 3, bg=fill(VC_HEX[v]))
        header_row(wv, ["LABORATÓRIO","META","FATURADO (arquivo)"])
        labs_v = [l for l in ordered if l in metas[v]]
        for i, lab in enumerate(labs_v, 3):
            bg = ALT if i%2==0 else None
            d = metas[v][lab]
            sty(wv.cell(row=i,column=1), lab, bg=bg, h="left")
            money(wv.cell(row=i,column=2), d["meta"] if d["meta"] else None, bg=bg)
            money(wv.cell(row=i,column=3), d["faturado"] if d["faturado"] else None, bg=bg)
        tr2 = 3+len(labs_v)
        for ci in range(1,4):
            c = wv.cell(row=tr2, column=ci); c.fill = HDR; c.border = tb()
            if ci == 1: sty(c, "TOTAL", bold=True, size=11, color="FFFFFF", bg=HDR)
            else: sum_col(wv, tr2, ci, 3)
        set_widths(wv, [26,16,16])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ── Interface ──────────────────────────────────────────────────────────────
st.markdown("""
<div class="hero">
    <h1>📊 Titãs · Relatórios 2026</h1>
    <p>Faça o upload dos arquivos abaixo e gere os relatórios com um clique.</p>
</div>
""", unsafe_allow_html=True)

col_up1, col_up2 = st.columns([1,1], gap="large")

with col_up1:
    st.markdown('<div class="section-label">Arquivos de configuração</div>', unsafe_allow_html=True)

    st.markdown('<div class="upload-card"><h4>📋 Meta mensal</h4><p>Arquivo com as metas por laboratório (ex: Abril.xlsx)</p></div>', unsafe_allow_html=True)
    f_meta = st.file_uploader("Meta", type=["xlsx"], label_visibility="collapsed", key="meta")

    st.markdown('<div class="upload-card"><h4>👥 Base Titãs</h4><p>Base de clientes por vendedora</p></div>', unsafe_allow_html=True)
    f_base = st.file_uploader("Base", type=["xlsx"], label_visibility="collapsed", key="base")

with col_up2:
    st.markdown('<div class="section-label">Dados de vendas por UF</div>', unsafe_allow_html=True)

    st.markdown('<div class="upload-card"><h4>🟢 Data — RS</h4><p>Export filtrado pela Filial 3 (AGROFARM RS)</p></div>', unsafe_allow_html=True)
    f_rs = st.file_uploader("RS", type=["xlsx"], label_visibility="collapsed", key="rs")

    st.markdown('<div class="upload-card"><h4>🔵 Data — RJ</h4><p>Export filtrado pelas Filiais 5 e 6 (AGROLIFE RJ)</p></div>', unsafe_allow_html=True)
    f_rj = st.file_uploader("RJ", type=["xlsx"], label_visibility="collapsed", key="rj")

    st.markdown('<div class="upload-card"><h4>🟣 Data — PR</h4><p>Export filtrado pela Filial 10 (AGROFARM PR)</p></div>', unsafe_allow_html=True)
    f_pr = st.file_uploader("PR", type=["xlsx"], label_visibility="collapsed", key="pr")

# Seleção do mês
st.markdown("---")
col_mes, col_btn = st.columns([1, 2], gap="large")
with col_mes:
    st.markdown('<div class="section-label">Mês de referência</div>', unsafe_allow_html=True)
    mes_idx = st.selectbox("Mês", MESES, index=3, label_visibility="collapsed")

# Status dos arquivos
arquivos_ok = {
    "Meta": f_meta is not None,
    "Base": f_base is not None,
    "RS":   f_rs   is not None,
    "RJ":   f_rj   is not None,
    "PR":   f_pr   is not None,
}
faltando = [k for k, v in arquivos_ok.items() if not v]

with col_btn:
    st.markdown('<div class="section-label">Status</div>', unsafe_allow_html=True)
    if faltando:
        st.warning(f"Aguardando: **{', '.join(faltando)}**")
    else:
        st.success("Todos os arquivos prontos!")

# Botão principal
st.markdown("---")
btn_col = st.columns([1,2,1])[1]
with btn_col:
    gerar = st.button("🚀 Gerar relatórios", disabled=bool(faltando),
                      use_container_width=True)

if gerar:
    with st.spinner("Processando..."):
        try:
            base_df, v_base = load_base(f_base.read())
            metas = load_meta(f_meta.read())
            files_uf = {}
            if f_rs: files_uf["RS"] = f_rs.read()
            if f_rj: files_uf["RJ"] = f_rj.read()
            if f_pr: files_uf["PR"] = f_pr.read()
            combined = load_all_data(files_uf, base_df)

            buf_mvr, summary, total_m, total_r = gerar_meta_vs_real(metas, combined, mes_idx)
            buf_cruz, n_cli = gerar_cruzamento(combined, mes_idx)
            buf_metas = gerar_metas_excel(metas, mes_idx)

            st.success("✅ Relatórios gerados!")

            # Métricas
            st.markdown(f'<div class="section-label" style="margin-top:1.5rem">Resultado — {mes_idx.capitalize()} 2026</div>', unsafe_allow_html=True)
            cols = st.columns(len(summary)+1)
            for i, (v, s) in enumerate(summary.items()):
                p = s["pct"]
                cor = "#1A6B3A" if p>=1.0 else ("#CC6600" if p>=0.85 else "#CC0000")
                with cols[i]:
                    st.markdown(f"""
                    <div class="metric-box">
                        <div class="metric-label">{v}</div>
                        <div class="metric-value" style="color:{cor}">{p:.1%}</div>
                        <div class="metric-sub">R$ {s['real']:,.0f} / R$ {s['meta']:,.0f}</div>
                    </div>""", unsafe_allow_html=True)
            with cols[-1]:
                gp = total_r/total_m if total_m else 0
                cor = "#1A6B3A" if gp>=1.0 else ("#CC6600" if gp>=0.85 else "#CC0000")
                st.markdown(f"""
                <div class="metric-box" style="border:2px solid #1F3864">
                    <div class="metric-label" style="color:#1F3864;font-weight:700">TOTAL TITÃS</div>
                    <div class="metric-value" style="color:{cor}">{gp:.1%}</div>
                    <div class="metric-sub">R$ {total_r:,.0f} / R$ {total_m:,.0f}</div>
                </div>""", unsafe_allow_html=True)

            # Downloads
            st.markdown("---")
            st.markdown('<div class="section-label">Downloads</div>', unsafe_allow_html=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M")
            d1, d2, d3 = st.columns(3)
            with d1:
                st.download_button("📥 Meta vs. Realizado", buf_mvr,
                    f"Meta_vs_Realizado_{mes_idx}_{ts}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
                st.caption("Meta × Realizado × GAP × % por laboratório")
            with d2:
                st.download_button("📥 Cruzamento de Vendas", buf_cruz,
                    f"Cruzamento_Titãs_{mes_idx}_{ts}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
                st.caption(f"{n_cli} clientes com vendas")
            with d3:
                st.download_button("📥 Metas Compiladas", buf_metas,
                    f"Metas_Titãs_{mes_idx}_{ts}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
                st.caption("Metas por laboratório e vendedora")

        except Exception as e:
            st.error(f"❌ Erro: {e}")
            import traceback; st.code(traceback.format_exc())

st.markdown("""
<div style="text-align:center;color:#AAB8C8;font-size:.78rem;margin-top:3rem">
    Titãs · Relatórios 2026 — os arquivos não são armazenados
</div>
""", unsafe_allow_html=True)
